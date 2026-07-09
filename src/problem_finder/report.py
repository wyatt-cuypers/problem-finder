import random
import time
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook

HEADERS = ["Rank", "Problem", "Composite score", "Mentions", "Threads",
           "Authors", "Subreddits", "First seen", "Last seen", "Spike?",
           "Unsolved-ness", "Solution landscape", "Examples", "Permalinks",
           "Opportunity", "Coherence flag"]

CLUSTER_SQL = """\
SELECT c.id, c.canonical_statement, c.solution_summary, c.opportunity_note,
       c.coherence_flag, s.*
FROM clusters c JOIN cluster_scores s ON s.cluster_id = c.id
WHERE s.passes_gate = 1
ORDER BY s.composite DESC"""

MEMBERS_SQL = """\
SELECT i.thread_id, i.subreddit, i.created_utc, i.score, i.permalink,
       e.problem_statement
FROM cluster_members cm
JOIN items i ON i.id = cm.item_id
JOIN extractions e ON e.item_id = cm.item_id
WHERE cm.cluster_id = ?"""


def _month(ts: int) -> str:
    return datetime.fromtimestamp(ts, tz=timezone.utc).strftime("%Y-%m")


def format_example(statement: str, subreddit: str, created_utc: int) -> str:
    return f"{statement} (r/{subreddit}, {_month(created_utc)})"


def pick_examples(members: list[dict], n: int = 3) -> list[dict]:
    best: dict[str, dict] = {}
    for m in members:
        cur = best.get(m["thread_id"])
        if cur is None or (m["score"] or 0) > (cur["score"] or 0):
            best[m["thread_id"]] = m
    ranked = sorted(best.values(), key=lambda m: m["score"] or 0, reverse=True)
    return ranked[:n]


def _cluster_row(rank: int, c, members: list[dict]) -> list:
    ex = pick_examples(members)
    return [rank, c["canonical_statement"], round(c["composite"], 3),
            c["mentions"], c["threads"], c["authors"], c["subreddits"],
            _month(c["first_seen"]), _month(c["last_seen"]),
            "yes" if c["spike_flag"] else "", round(c["unsolvedness"], 2),
            c["solution_summary"],
            " | ".join(format_example(m["problem_statement"], m["subreddit"],
                                      m["created_utc"]) for m in ex),
            " | ".join(f"https://reddit.com{m['permalink']}" for m in ex),
            c["opportunity_note"],
            "check: may be merged topics" if c["coherence_flag"] else ""]


def _write_sheet(ws, clusters, conn, limit=None):
    ws.append(HEADERS)
    for rank, c in enumerate(clusters[:limit], start=1):
        members = [dict(r) for r in conn.execute(MEMBERS_SQL, (c["id"],))]
        ws.append(_cluster_row(rank, c, members))


def _run_info(ws, conn, usage):
    ws.append(["Subreddit", "Run time", "Oldest", "Newest", "Posts", "Comments"])
    for r in conn.execute("SELECT * FROM coverage ORDER BY subreddit, run_ts"):
        ws.append([r["subreddit"], _month(r["run_ts"]),
                   _month(r["oldest_utc"]) if r["oldest_utc"] else "",
                   _month(r["newest_utc"]) if r["newest_utc"] else "",
                   r["post_count"], r["comment_count"]])
    n_failed = conn.execute(
        "SELECT COUNT(*) AS n FROM extractions WHERE status='failed'"
    ).fetchone()["n"]
    ws.append([])
    ws.append(["Failed extractions", n_failed])
    if usage:
        ws.append(["Gemini calls", usage["calls"]])
        ws.append(["Input tokens", usage["input_tokens"]])
        ws.append(["Output tokens", usage["output_tokens"]])


def _extraction_qa(ws, conn, sample_size=30):
    rows = conn.execute("""\
        SELECT i.text, e.is_problem, e.problem_statement, e.solution_mentioned
        FROM extractions e JOIN items i ON i.id = e.item_id
        WHERE e.status='ok'""").fetchall()
    random.seed(42)
    sample = random.sample(rows, min(sample_size, len(rows)))
    ws.append(["Raw text (truncated)", "Is problem?", "Extracted statement",
               "Solution mentioned"])
    for r in sample:
        ws.append([r["text"][:500], bool(r["is_problem"]),
                   r["problem_statement"], r["solution_mentioned"]])


def _markdown(clusters, conn) -> str:
    lines = ["# Top 10 problem opportunities\n"]
    for rank, c in enumerate(clusters[:10], start=1):
        members = [dict(r) for r in conn.execute(MEMBERS_SQL, (c["id"],))]
        ex = pick_examples(members)
        lines += [
            f"## {rank}. {c['canonical_statement']}\n",
            f"- **Score:** {c['composite']:.2f} | **Mentions:** "
            f"{c['mentions']} across {c['threads']} threads, "
            f"{c['authors']} authors, {c['subreddits']} subreddits",
            f"- **Seen:** {_month(c['first_seen'])} to {_month(c['last_seen'])}"
            + (" (possible news spike)" if c["spike_flag"] else ""),
            f"- **Unsolved-ness:** {c['unsolvedness']:.2f}",
            f"- **Current solutions:** {c['solution_summary']}",
            f"- **Opportunity:** {c['opportunity_note']}",
            "- **Examples:**",
            *[f"  - {format_example(m['problem_statement'], m['subreddit'], m['created_utc'])}"
              for m in ex],
            "",
        ]
    return "\n".join(lines)


def run_report(conn, cfg, usage: dict | None = None) -> tuple[Path, Path]:
    out = Path(cfg.output_dir)
    out.mkdir(parents=True, exist_ok=True)
    ts = time.strftime("%Y%m%d_%H%M%S")
    clusters = [dict(r) for r in conn.execute(CLUSTER_SQL)]

    wb = Workbook()
    _write_sheet(wb.active, clusters, conn, limit=50)
    wb.active.title = "Top 50"
    _write_sheet(wb.create_sheet("All clusters"), clusters, conn)
    _run_info(wb.create_sheet("Run info"), conn, usage)
    if cfg.pilot:
        _extraction_qa(wb.create_sheet("Extraction QA"), conn)
    xlsx_path = out / f"report_{ts}.xlsx"
    wb.save(xlsx_path)

    md_path = out / f"summary_{ts}.md"
    md_path.write_text(_markdown(clusters, conn))
    print(f"report: {xlsx_path} and {md_path} "
          f"({len(clusters)} clusters pass gates)")
    return xlsx_path, md_path
