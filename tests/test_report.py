from datetime import datetime, timezone

from openpyxl import load_workbook

from problem_finder import db
from problem_finder.report import format_example, pick_examples, run_report
from tests.test_db import make_item

END = 1_750_000_000


def test_format_example():
    ts = int(datetime(2026, 3, 15, tzinfo=timezone.utc).timestamp())
    assert format_example("hotel wifi is unreliable", "travel", ts) == \
        "hotel wifi is unreliable (r/travel, 2026-03)"


def test_pick_examples_distinct_threads_by_score():
    members = [
        dict(thread_id="t1", score=50, problem_statement="a",
             subreddit="s", created_utc=END),
        dict(thread_id="t1", score=99, problem_statement="b",
             subreddit="s", created_utc=END),
        dict(thread_id="t2", score=10, problem_statement="c",
             subreddit="s", created_utc=END),
    ]
    picked = pick_examples(members, n=3)
    assert [m["problem_statement"] for m in picked] == ["b", "c"]


def _seed(conn):
    items = [make_item(f"i{k}", thread_id=f"t{k}", author_hash=f"a{k}",
                       created_utc=END - k * 86400, score=10 + k,
                       subreddit="running" if k % 2 else "travel")
             for k in range(4)]
    db.upsert_items(conn, items)
    for k in range(4):
        conn.execute("INSERT INTO extractions VALUES (?,?,?,?,?,?,?,?,?)",
                     (f"i{k}", "ok", 1, f"statement {k}", "other", "none",
                      None, 0, 1))
    conn.execute("""INSERT INTO clusters VALUES
        (1, 'canonical problem', 'no real solutions', 'big market', 0)""")
    conn.executemany("INSERT INTO cluster_members VALUES (?,1)",
                     [(f"i{k}",) for k in range(4)])
    conn.execute("""INSERT INTO cluster_scores VALUES
        (1, 4, 4, 4, 2, ?, ?, 0, 0.9, 1.45, 1)""", (END - 3 * 86400, END))
    conn.commit()


def test_run_report_writes_xlsx_and_md(tmp_path):
    conn = db.connect(":memory:")
    _seed(conn)

    class Cfg:
        output_dir = str(tmp_path)
        pilot = True
        window_days = 14

    xlsx, md = run_report(conn, Cfg(), usage={"calls": 3, "input_tokens": 10,
                                              "output_tokens": 5})
    wb = load_workbook(xlsx)
    assert {"Top 50", "All clusters", "Run info",
            "Extraction QA"} <= set(wb.sheetnames)
    ws = wb["Top 50"]
    header = [c.value for c in ws[1]]
    assert "Problem" in header and "Composite score" in header
    assert ws.cell(row=2, column=header.index("Problem") + 1).value == \
        "canonical problem"
    text = md.read_text()
    assert "canonical problem" in text
    assert "big market" in text
